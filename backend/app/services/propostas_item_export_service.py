from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_LEVEL_1_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
HEADER_LEVEL_2_FILL = PatternFill(fill_type="solid", fgColor="2E75B6")
WHITE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
GREEN_FILL = PatternFill(fill_type="solid", fgColor="00B050")

THIN_GRAY = Side(style="thin", color="AAAAAA")
THIN_LIGHT = Side(style="thin", color="CCCCCC")
MEDIUM_GREEN = Side(style="medium", color="1A7F3C")

HEADER_LEVEL_1_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
HEADER_LEVEL_2_FONT = Font(bold=True, color="FFFFFF", size=10, name="Arial")
TEXT_FONT = Font(size=10, name="Arial")
GREEN_FONT = Font(bold=True, color="1A4731", size=10, name="Arial")

ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


FIXED_COLUMNS = [
    ("numero_item", "Item", 6),
    ("descricao", "Descricao", 22),
    ("descricao_detalhada", "Descricao Detalhada", 45),
    ("quantidade_solicitada", "Qtde Solicitada", 10),
    ("valor_estimado_unitario", "Valor Estimado Unitario", 18),
]

COMPANY_COLUMNS = [
    ("cnpj", "CNPJ", 20),
    ("nome_empresa", "Nome da Empresa", 40),
    ("valor_unitario_ofertado", "Valor Unitario Ofertado", 20),
]


def build_propostas_item_workbook(
    *,
    portal_sigla: str,
    numero_processo: str,
    items: list[dict[str, object]],
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Propostas"
    sheet.freeze_panes = "A3"

    max_propostas = max((len(_as_list(item.get("propostas"))) for item in items), default=0)

    fixed_column_count = len(FIXED_COLUMNS)
    company_column_count = len(COMPANY_COLUMNS)
    current_col = 1

    for _, label, width in FIXED_COLUMNS:
        cell = sheet.cell(row=1, column=current_col, value=label)
        _apply_header_level_1(cell)
        sheet.merge_cells(start_row=1, start_column=current_col, end_row=2, end_column=current_col)
        sheet.column_dimensions[get_column_letter(current_col)].width = width
        current_col += 1

    for company_index in range(max_propostas):
        start_col = current_col + company_index * company_column_count
        end_col = start_col + company_column_count - 1
        header_cell = sheet.cell(row=1, column=start_col, value=f"Empresa {company_index + 1}")
        _apply_header_level_1(header_cell)
        if start_col != end_col:
            sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        for offset, (_, label, width) in enumerate(COMPANY_COLUMNS):
            subcell = sheet.cell(row=2, column=start_col + offset, value=label)
            _apply_header_level_2(subcell)
            sheet.column_dimensions[get_column_letter(start_col + offset)].width = width

    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 28

    for row_index, item in enumerate(items, start=3):
        sheet.row_dimensions[row_index].height = 40
        lowest_index = _find_lowest_proposal_index(_as_list(item.get("propostas")))

        for column_index, (key, _, _) in enumerate(FIXED_COLUMNS, start=1):
            raw_value = item.get(key)
            cell = sheet.cell(row=row_index, column=column_index, value=_coerce_cell_value(key, raw_value))
            _apply_text_cell(cell)
            if key == "valor_estimado_unitario":
                _apply_monetary_if_possible(cell, raw_value)

        for company_index in range(max_propostas):
            start_col = fixed_column_count + 1 + company_index * company_column_count
            proposal = _as_list(item.get("propostas"))[company_index] if company_index < len(_as_list(item.get("propostas"))) else {}

            for offset, (field, _, _) in enumerate(COMPANY_COLUMNS):
                raw_value = proposal.get(field, "") if isinstance(proposal, dict) else ""
                cell = sheet.cell(
                    row=row_index,
                    column=start_col + offset,
                    value=_coerce_cell_value(field, raw_value),
                )

                is_lowest = lowest_index == company_index
                if field == "valor_unitario_ofertado":
                    _apply_monetary_if_possible(cell, raw_value, is_lowest=is_lowest)
                else:
                    _apply_text_cell(cell, highlight=is_lowest)

    metadata_sheet = workbook.create_sheet("Resumo")
    metadata_sheet["A1"] = "Portal"
    metadata_sheet["B1"] = portal_sigla
    metadata_sheet["A2"] = "Processo"
    metadata_sheet["B2"] = numero_processo
    metadata_sheet["A3"] = "Itens extraidos"
    metadata_sheet["B3"] = len(items)
    metadata_sheet["A4"] = "Maximo de propostas por item"
    metadata_sheet["B4"] = max_propostas
    metadata_sheet.column_dimensions["A"].width = 28
    metadata_sheet.column_dimensions["B"].width = 48

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_propostas_item_filename(portal_sigla: str, numero_processo: str) -> str:
    portal_part = _sanitize_filename_part(portal_sigla or "portal")
    process_part = _sanitize_filename_part(numero_processo or "licitacao")
    return f"propostas_{portal_part}_{process_part}.xlsx"


def _apply_header_level_1(cell) -> None:
    cell.font = HEADER_LEVEL_1_FONT
    cell.fill = HEADER_LEVEL_1_FILL
    cell.alignment = ALIGN_CENTER_WRAP
    cell.border = Border(top=THIN_GRAY, bottom=THIN_GRAY, left=THIN_GRAY, right=THIN_GRAY)


def _apply_header_level_2(cell) -> None:
    cell.font = HEADER_LEVEL_2_FONT
    cell.fill = HEADER_LEVEL_2_FILL
    cell.alignment = ALIGN_CENTER_WRAP
    cell.border = Border(top=THIN_GRAY, bottom=THIN_GRAY, left=THIN_GRAY, right=THIN_GRAY)


def _apply_text_cell(cell, *, highlight: bool = False) -> None:
    cell.font = GREEN_FONT if highlight else TEXT_FONT
    cell.fill = GREEN_FILL if highlight else WHITE_FILL
    cell.alignment = ALIGN_LEFT_WRAP
    border_side = MEDIUM_GREEN if highlight else THIN_LIGHT
    cell.border = Border(top=border_side, bottom=border_side, left=border_side, right=border_side)


def _apply_monetary_if_possible(cell, raw_value: object, *, is_lowest: bool = False) -> None:
    parsed = _parse_brl(raw_value)
    if parsed is None:
        _apply_text_cell(cell, highlight=is_lowest)
        cell.alignment = ALIGN_RIGHT
        return

    cell.value = parsed
    cell.number_format = '"R$" #,##0.0000'
    cell.font = GREEN_FONT if is_lowest else TEXT_FONT
    cell.fill = GREEN_FILL if is_lowest else WHITE_FILL
    cell.alignment = ALIGN_RIGHT
    border_side = MEDIUM_GREEN if is_lowest else THIN_LIGHT
    cell.border = Border(top=border_side, bottom=border_side, left=border_side, right=border_side)


def _coerce_cell_value(field: str, value: object) -> object:
    text = _stringify(value)
    if field in {"descricao", "descricao_detalhada"}:
        return text
    if not text:
        return ""
    return text


def _find_lowest_proposal_index(propostas: list[dict[str, object]]) -> int | None:
    best_index: int | None = None
    best_value: float | None = None
    for index, proposta in enumerate(propostas):
        parsed = _parse_brl(proposta.get("valor_unitario_ofertado"))
        if parsed is None:
            continue
        if best_value is None or parsed < best_value:
            best_value = parsed
            best_index = index
    return best_index


def _parse_brl(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text.upper() == "[NAO INFORMADO]":
        return None
    text = text.replace("R$", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _sanitize_filename_part(value: str) -> str:
    sanitized = "".join(char if char.isalnum() else "_" for char in value.strip().lower())
    sanitized = sanitized.strip("_")
    return sanitized or "arquivo"


def _stringify(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.4f}".rstrip("0").rstrip(".")
    return str(value)


def _as_list(value: object) -> list[dict[str, object]]:
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    return []
